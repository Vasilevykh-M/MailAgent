from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from mail_agent.config import TableSettings
from mail_agent.storage.workbook import WorkbookRepository


class MemoryDrive:
    def __init__(self, tmp_path: Path) -> None:
        self.path = tmp_path / "remote.xlsx"
        self.path.parent.mkdir(exist_ok=True)
        self.results: dict[str, bytes] = {}

    def metadata(self, remote_path: str):
        if not self.path.exists():
            raise FileNotFoundError(remote_path)
        return {
            "size": self.path.stat().st_size,
            "modified": str(self.path.stat().st_mtime),
            "md5": None,
            "sha256": None,
        }

    def download(self, remote_path: str, destination: Path):
        destination.write_bytes(self.path.read_bytes())
        return destination

    def upload(self, local_path: Path, remote_path: str, *, overwrite: bool):
        self.path.write_bytes(local_path.read_bytes())
        return self.metadata(remote_path)

    def upload_bytes(self, data: bytes, remote_path: str, *, overwrite: bool):
        self.results[remote_path] = data
        return {"path": remote_path}


def _result(record_id: str) -> dict:
    return {
        "record_id": record_id,
        "mailbox": "INBOX",
        "uid": "1",
        "message_id": "<1>",
        "message": {
            "date": "2026-07-12T10:00:00+00:00",
            "from": "sender@example.test",
            "to": ["recipient@example.test"],
            "cc": [],
            "bcc": [],
            "reply_to": [],
            "subject": "subject",
            "flags": ["\\Seen"],
        },
        "body": "body",
        "attachments": [
            {
                "original_filename": "document.pdf",
                "status": "processed",
                "processing_tool": "ocr",
                "language": "ru",
                "warnings": ["Низкая уверенность распознавания."],
                "raw_extracted_text": "Текст из вложения",
            }
        ],
        "summary": {
            "summary_ru": "итог",
            "attachment_summaries": ["Вложение обработано."],
            "key_facts_ru": ["Факт"],
            "action_items_ru": ["Действие"],
            "deadlines": ["2026-07-20"],
            "warnings_ru": ["Предупреждение"],
            "confidence": 0.9,
        },
    }


def test_workbook_appends_then_updates_same_record(tmp_path: Path) -> None:
    drive = MemoryDrive(tmp_path)
    settings = TableSettings(create_if_missing=True, report_directory="/results")
    storage = WorkbookRepository(drive, settings)
    storage.upsert(_result("same"))
    changed = _result("same")
    changed["summary"]["summary_ru"] = "обновлено"
    storage.upsert(changed)
    book = load_workbook(drive.path)
    sheet = book[settings.sheet_name]
    headers = {cell.value: cell.column for cell in sheet[settings.header_row]}
    assert [cell.value for cell in sheet[settings.header_row]] == [
        "Отправитель",
        "Дата письма",
        "Тема",
        "Итоговая суммаризация",
        "Суммаризация вложений",
        "Ключевые факты и особенности",
        "ID записи",
    ]
    assert sheet.cell(2, headers["ID записи"]).value == "same"
    assert sheet.column_dimensions["G"].hidden is True
    assert sheet.freeze_panes == "A2"
    assert sheet.cell(2, headers["Итоговая суммаризация"]).alignment.wrap_text is True
    assert sheet.cell(2, headers["Итоговая суммаризация"]).alignment.vertical == "top"
    assert sheet.cell(2, headers["Итоговая суммаризация"]).value == "обновлено"
    assert sheet.cell(2, headers["Ключевые факты и особенности"]).value == "• Факт\nОсобенности:\n• Предупреждение"
    report = drive.results["/results/same.md"].decode("utf-8")
    assert "## Итоговая суммаризация" in report
    assert "## Содержимое письма" in report
    assert "### Извлечённый текст из вложения 1: document.pdf" in report
    assert '"summary_ru"' not in report
    assert '"attachment_summaries"' not in report
    book.close()


def test_workbook_migrates_legacy_columns_to_compact_layout(tmp_path: Path) -> None:
    drive = MemoryDrive(tmp_path)
    book = Workbook()
    book.active.title = "Письма"
    book.active.cell(1, 1, "ID записи")
    book.active.cell(1, 2, "Отправитель")
    book.active.cell(1, 3, "Ключевые факты")
    book.active.cell(1, 4, "Предупреждения")
    book.active.append(["legacy", "sender@example.test", "• Факт", "• Особенность"])
    book.save(drive.path)
    book.close()
    storage = WorkbookRepository(drive, TableSettings(create_if_missing=True))

    storage.upsert(_result("renamed"))

    uploaded = load_workbook(drive.path)
    headers = [cell.value for cell in uploaded["Письма"][1]]
    assert headers == list(storage.settings.columns.values())
    assert uploaded["Письма"].cell(2, 6).value == "• Факт\nОсобенности:\n• Особенность"
    uploaded.close()
