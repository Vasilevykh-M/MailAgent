"""Сериализованное optimistic-обновление .xlsx на Яндекс Диске."""

from __future__ import annotations

import logging
import tempfile
import time
from pathlib import Path
from threading import Lock
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ..config import TableSettings
from ..exceptions import WorkbookConflictError, WorkbookMissingError
from ..integrations.drive import DriveGateway
from ..logging import log_event

EXCEL_CELL_LIMIT = 32_767
LOGGER = logging.getLogger(__name__)
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_COLUMN_WIDTHS = {
    "Отправитель": 28,
    "Дата письма": 22,
    "Тема": 38,
    "Итоговая суммаризация": 52,
    "Суммаризация вложений": 42,
    "Ключевые факты и особенности": 42,
}


class WorkbookRepository:
    """Не пересоздаёт существующую книгу и блокирует обновления в пределах процесса."""

    def __init__(self, drive: DriveGateway, settings: TableSettings) -> None:
        self.drive, self.settings = drive, settings
        self._lock = Lock()

    @staticmethod
    def _identity(metadata: dict[str, Any]) -> tuple[Any, ...]:
        return tuple(metadata.get(name) for name in ("md5", "sha256", "modified", "size"))

    @staticmethod
    def _text(value: object, *, empty: str = "—") -> str:
        if value is None:
            return empty
        rendered = str(value).strip()
        return rendered or empty

    @classmethod
    def _items(cls, value: object) -> list[str]:
        if isinstance(value, (list, tuple, set)):
            return [cls._text(item, empty="") for item in value if cls._text(item, empty="")]
        item = cls._text(value, empty="")
        return [item] if item else []

    @classmethod
    def _bullets(cls, value: object) -> str:
        items = cls._items(value)
        return "\n".join(f"• {item}" for item in items) if items else "—"

    @classmethod
    def _recipients(cls, message: dict[str, Any]) -> str:
        labels = (("Кому", "to"), ("Копия", "cc"), ("Скрытая копия", "bcc"), ("Ответить", "reply_to"))
        lines = [
            f"{label}: {', '.join(cls._items(message.get(key)))}"
            for label, key in labels
            if cls._items(message.get(key))
        ]
        return "\n".join(lines) if lines else "—"

    @classmethod
    def _attachments_cell(cls, attachments: object) -> str:
        if not isinstance(attachments, list) or not attachments:
            return "—"
        rows: list[str] = []
        for index, item in enumerate(attachments, 1):
            if not isinstance(item, dict):
                continue
            name = cls._text(item.get("original_filename"), empty="Без имени")
            status = cls._text(item.get("status"))
            tool = cls._text(item.get("processing_tool"))
            row = f"{index}. {name}\n   Статус: {status}; способ: {tool}"
            language = cls._text(item.get("language"), empty="")
            if language:
                row += f"; язык: {language}"
            warnings = cls._items(item.get("warnings"))
            if warnings:
                row += "\n   Предупреждения: " + "; ".join(warnings)
            rows.append(row)
        return "\n\n".join(rows) if rows else "—"

    def _result_report_path(self, record_id: str) -> str:
        return self.settings.report_directory.rstrip("/") + f"/{record_id}.md"

    def _markdown_report(self, result: dict[str, Any]) -> str:
        message_value = result.get("message")
        summary_value = result.get("summary")
        attachments_value = result.get("attachments")
        message: dict[str, Any] = message_value if isinstance(message_value, dict) else {}
        summary: dict[str, Any] = summary_value if isinstance(summary_value, dict) else {}
        attachments: list[Any] = attachments_value if isinstance(attachments_value, list) else []
        lines = [
            "# Отчёт по письму",
            "",
            "## Реквизиты",
            f"- **ID записи:** {self._text(result.get('record_id'))}",
            f"- **Папка:** {self._text(result.get('mailbox'))}",
            f"- **UID:** {self._text(result.get('uid'))}",
            f"- **Дата письма:** {self._text(message.get('date'))}",
            f"- **Отправитель:** {self._text(message.get('from'))}",
            f"- **Получатели:** {self._recipients(message)}",
            f"- **Тема:** {self._text(message.get('subject'))}",
            "",
            "## Итоговая суммаризация",
            self._text(summary.get("summary_ru")),
            "",
            "## Ключевые факты",
            self._bullets(summary.get("key_facts_ru")),
            "",
            "## Необходимые действия",
            self._bullets(summary.get("action_items_ru")),
            "",
            "## Сроки",
            self._bullets(summary.get("deadlines")),
            "",
            "## Вложения",
            self._attachments_cell(attachments),
        ]
        if message.get("is_forwarded"):
            lines.insert(10, f"- **Переслал:** {self._text(message.get('forwarded_by'))}")
        for index, attachment in enumerate(attachments, 1):
            if not isinstance(attachment, dict):
                continue
            extracted = attachment.get("corrected_text") or attachment.get("raw_extracted_text")
            if not isinstance(extracted, str) or not extracted.strip():
                continue
            lines.extend(
                [
                    "",
                    f"### Извлечённый текст из вложения {index}: {self._text(attachment.get('original_filename'))}",
                    extracted.strip(),
                ]
            )
        body = result.get("body")
        if isinstance(body, str) and body.strip():
            lines.extend(["", "## Содержимое письма", body.strip()])
        warnings = self._bullets(summary.get("warnings_ru"))
        if warnings != "—":
            lines.extend(["", "## Предупреждения", warnings])
        return "\n".join(lines).rstrip() + "\n"

    def _save_full_report(self, result: dict[str, Any]) -> str | None:
        if not self.settings.save_full_report:
            return None
        remote = self._result_report_path(str(result["record_id"]))
        self.drive.upload_bytes(self._markdown_report(result).encode("utf-8"), remote, overwrite=True)
        return remote

    def _book(self, path: Path) -> Workbook:
        if path.exists():
            return load_workbook(path, data_only=False)
        if not self.settings.create_if_missing:
            raise WorkbookMissingError("Файл Excel не найден, а create_if_missing=false.")
        workbook = Workbook()
        workbook.active.title = self.settings.sheet_name
        return workbook

    @staticmethod
    def _value(sheet: Any, row: int, headers: dict[str, int], header: str) -> object:
        column = headers.get(header)
        return sheet.cell(row, column).value if column is not None else None

    @classmethod
    def _combined_facts_and_features(cls, facts: object, features: object) -> str:
        facts_text = cls._text(facts, empty="")
        features_text = cls._text(features, empty="")
        if not features_text:
            return facts_text or "—"
        if not facts_text:
            return f"Особенности:\n{features_text}"
        return f"{facts_text}\nОсобенности:\n{features_text}"

    def _existing_rows(self, sheet: Any, headers: dict[str, int]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for row in range(self.settings.header_row + 1, sheet.max_row + 1):
            values = {
                "sender": self._value(sheet, row, headers, "Отправитель"),
                "message_date": self._value(sheet, row, headers, "Дата письма"),
                "subject": self._value(sheet, row, headers, "Тема"),
                "summary": self._value(sheet, row, headers, "Итоговая суммаризация"),
                "attachment_summary": self._value(sheet, row, headers, "Суммаризация вложений"),
                "key_facts": self._combined_facts_and_features(
                    self._value(sheet, row, headers, "Ключевые факты и особенности")
                    or self._value(sheet, row, headers, "Ключевые факты"),
                    self._value(sheet, row, headers, "Предупреждения"),
                ),
                "record_id": self._value(sheet, row, headers, "ID записи"),
            }
            if any(value not in (None, "", "—") for value in values.values()):
                rows.append(values)
        return rows

    def _style_sheet(self, sheet: Any) -> None:
        sheet.freeze_panes = f"A{self.settings.header_row + 1}"
        sheet.auto_filter.ref = f"A{self.settings.header_row}:{chr(64 + len(self.settings.columns))}{sheet.max_row}"
        for column, title in enumerate(self.settings.columns.values(), 1):
            cell = sheet.cell(self.settings.header_row, column)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.column_dimensions[cell.column_letter].width = _COLUMN_WIDTHS.get(title, 2)
            if title == "ID записи":
                sheet.column_dimensions[cell.column_letter].hidden = True
        for row in range(self.settings.header_row + 1, sheet.max_row + 1):
            self._style_data_row(sheet, row)

    def _style_data_row(self, sheet: Any, row: int) -> None:
        for column in range(1, len(self.settings.columns) + 1):
            sheet.cell(row, column).alignment = Alignment(vertical="top", wrap_text=True)

    def _headers(self, sheet: Any) -> dict[str, int]:
        titles = list(self.settings.columns.values())
        existing = {cell.value: cell.column for cell in sheet[self.settings.header_row] if isinstance(cell.value, str)}
        current_titles = [sheet.cell(self.settings.header_row, column).value for column in range(1, len(titles) + 1)]
        if current_titles != titles or sheet.max_column != len(titles):
            previous_rows = self._existing_rows(sheet, existing)
            if sheet.max_row:
                sheet.delete_rows(1, sheet.max_row)
            if sheet.max_column:
                sheet.delete_cols(1, sheet.max_column)
            for column, title in enumerate(titles, 1):
                sheet.cell(self.settings.header_row, column, title)
            for row, values in enumerate(previous_rows, self.settings.header_row + 1):
                for key, title in self.settings.columns.items():
                    sheet.cell(row, titles.index(title) + 1, values.get(key))
        self._style_sheet(sheet)
        return {title: index for index, title in enumerate(titles, 1)}

    def _prepare_values(self, result: dict[str, Any], full_path: str | None) -> dict[str, object]:
        summary = result["summary"]
        values: dict[str, object] = {
            "sender": result["message"].get("from"),
            "message_date": result["message"].get("date"),
            "subject": result["message"].get("subject"),
            "summary": summary.get("summary_ru"),
            "attachment_summary": self._bullets(summary.get("attachment_summaries")),
            "key_facts": self._combined_facts_and_features(
                self._bullets(summary.get("key_facts_ru")), self._bullets(summary.get("warnings_ru"))
            ),
            "record_id": result["record_id"],
        }
        return values

    def _apply(self, file_path: Path, result: dict[str, Any], full_path: str | None) -> None:
        book = self._book(file_path)
        try:
            sheet = (
                book[self.settings.sheet_name]
                if self.settings.sheet_name in book.sheetnames
                else book.create_sheet(self.settings.sheet_name)
            )
            headers = self._headers(sheet)
            record_column = headers[self.settings.columns["record_id"]]
            row = next(
                (
                    index
                    for index in range(self.settings.header_row + 1, sheet.max_row + 1)
                    if sheet.cell(index, record_column).value == result["record_id"]
                ),
                None,
            )
            if row is None:
                row = max(sheet.max_row + 1, self.settings.header_row + 1)
            values = self._prepare_values(result, full_path)
            oversized = False
            for key, title in self.settings.columns.items():
                value = values.get(key)
                if isinstance(value, str) and len(value) > EXCEL_CELL_LIMIT:
                    oversized = True
                    value = (
                        value[: EXCEL_CELL_LIMIT - 90] + "\n[Текст сокращён; полная версия хранится в Markdown-отчёте.]"
                    )
                sheet.cell(row, headers[title], value)
            self._style_data_row(sheet, row)
            if oversized and not full_path:
                raise WorkbookMissingError("Значение не помещается в Excel, а сохранение полного отчёта отключено.")
            book.save(file_path)
        finally:
            book.close()

    def _verify_row(self, path: Path, record_id: str) -> bool:
        book = load_workbook(path, read_only=True, data_only=False)
        try:
            if self.settings.sheet_name not in book.sheetnames:
                return False
            sheet = book[self.settings.sheet_name]
            headers = {
                cell.value: cell.column for cell in sheet[self.settings.header_row] if isinstance(cell.value, str)
            }
            col = headers.get(self.settings.columns["record_id"])
            if col is None:
                return False
            return any(
                sheet.cell(row, col).value == record_id
                for row in range(self.settings.header_row + 1, sheet.max_row + 1)
            )
        finally:
            book.close()

    def upsert(self, result: dict[str, Any]) -> dict[str, Any]:
        """Загружает, проверяет конфликт, обновляет, загружает и повторно скачивает книгу."""

        started = time.perf_counter()
        record = str(result.get("record_id", ""))
        log_event(LOGGER, "workbook_update_started", component="workbook", operation="upsert", record_id=record)
        with self._lock, tempfile.TemporaryDirectory(prefix="mail-agent-xlsx-") as directory:
            temp = Path(directory)
            full_path = self._save_full_report(result)
            for attempt in range(self.settings.max_conflict_retries + 1):
                try:
                    original = self.drive.metadata(self.settings.remote_path)
                    exists = True
                except Exception as exc:
                    original, exists = {}, False
                    if not self.settings.create_if_missing:
                        raise WorkbookMissingError("Файл Excel на Яндекс Диске не найден.") from exc
                source = temp / "source.xlsx"
                if exists:
                    self.drive.download(self.settings.remote_path, source)
                self._apply(source, result, full_path)
                current = self.drive.metadata(self.settings.remote_path) if exists else {}
                if exists and self._identity(current) != self._identity(original):
                    log_event(
                        LOGGER,
                        "workbook_update_conflict",
                        level=logging.WARNING,
                        component="workbook",
                        operation="upsert",
                        record_id=record,
                        attempt=attempt + 1,
                        max_attempts=self.settings.max_conflict_retries + 1,
                        retryable=attempt < self.settings.max_conflict_retries,
                    )
                    if attempt == self.settings.max_conflict_retries:
                        raise WorkbookConflictError("Файл Excel изменился во время обновления.")
                    continue
                uploaded = self.drive.upload(source, self.settings.remote_path, overwrite=True)
                verify = temp / "verify.xlsx"
                self.drive.download(self.settings.remote_path, verify)
                if not self._verify_row(verify, str(result["record_id"])):
                    raise WorkbookConflictError("После загрузки не найдена строка record_id.")
                log_event(
                    LOGGER,
                    "workbook_update_completed",
                    component="workbook",
                    operation="upsert",
                    record_id=record,
                    attempt=attempt + 1,
                    max_attempts=self.settings.max_conflict_retries + 1,
                    duration_ms=round((time.perf_counter() - started) * 1000),
                )
                return {"remote_path": self.settings.remote_path, "metadata": uploaded, "full_result_path": full_path}
        raise WorkbookConflictError("Исчерпаны попытки обновить Excel.")
